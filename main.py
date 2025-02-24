import openpyxl
from docx import Document
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
from datetime import datetime
import os
import shutil

def format_date(date_str):
    months = {
        "01": "января", "02": "февраля", "03": "марта", "04": "апреля",
        "05": "мая", "06": "июня", "07": "июля", "08": "августа",
        "09": "сентября", "10": "октября", "11": "ноября", "12": "декабря"
    }
    
    day, month, year = date_str.split(".")
    return f"«{day}» {months[month]} {year}"

def get_cell_value(ws, cell_row, cell_column):
    cell = f"{cell_column}{cell_row}"
    return str(ws[cell].value if ws[cell].value is not None else "")

# Функция выбора файла Excel
def select_excel_file():
    global wb, excel_file_name, excel_file_path
    file_path = filedialog.askopenfilename(
        title="Выберите файл Excel",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        initialdir=os.path.join(os.getcwd(), "Excel_files")
    )
    if file_path:
        try:
            wb = openpyxl.load_workbook(file_path)
            excel_file_name = os.path.splitext(os.path.basename(file_path))[0]
            excel_file_path = file_path  # Сохраняем полный путь для сохранения
            load_ui_data(wb)
            messagebox.showinfo("Успех", f"Загружен файл: {file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить файл: {str(e)}")
    else:
        messagebox.showwarning("Предупреждение", "Файл не выбран!")

# Функция создания нового проекта
def create_project():
    global wb, excel_file_name, excel_file_path
    project_name = simpledialog.askstring("Создать проект", "Введите название проекта:")
    if project_name:
        template_path = os.path.join(os.getcwd(), "Templates/Excel_templates", "Template.xlsx")
        new_file_path = os.path.join(os.getcwd(), "Excel_files", f"{project_name}.xlsx")
        
        try:
            if not os.path.exists(template_path):
                messagebox.showerror("Ошибка", "Файл Template.xlsx не найден в папке Templates!")
                return
            
            shutil.copyfile(template_path, new_file_path)
            wb = openpyxl.load_workbook(new_file_path)
            excel_file_name = project_name
            excel_file_path = new_file_path  # Сохраняем путь к новому файлу
            load_ui_data(wb)
            messagebox.showinfo("Успех", f"Проект '{project_name}' создан и загружен: {new_file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать проект: {str(e)}")
    else:
        messagebox.showwarning("Предупреждение", "Название проекта не введено!")

# Функция сохранения данных в Excel
def save_to_excel():
    global wb, excel_file_path
    if wb is None:
        messagebox.showwarning("Предупреждение", "Сначала загрузите файл Excel!")
        return

    try:
        main_sheet = wb['Main data']
        # Обновляем данные в столбце B на основе полей ввода
        for _, value_data in main_data:
            cell = f"{value_data['cell_column']}{value_data['cell_row']}"
            main_sheet[cell].value = entries[value_data['name']].get()
        
        wb.save(excel_file_path)
        messagebox.showinfo("Успех", f"Данные сохранены в {excel_file_path}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {str(e)}")

# Функция загрузки данных в интерфейс
def load_ui_data(workbook):
    main_sheet = workbook['Main data']
    
    for widget in root.grid_slaves():
        if widget.grid_info().get('row', 0) in range(1, 11):
            widget.destroy()

    global entries
    entries = {}
    for i, (label_data, value_data) in enumerate(main_data, start=1):
        label_text = get_cell_value(main_sheet, label_data['cell_row'], 'A')
        value = get_cell_value(main_sheet, value_data['cell_row'], 'B')
        
        tk.Label(root, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky="e")
        entries[value_data['name']] = tk.Entry(root, width=100)
        entries[value_data['name']].grid(row=i, column=1, padx=5, pady=5)
        entries[value_data['name']].insert(0, value)

    for widget in scrollable_frame.winfo_children():
        widget.destroy()

    tk.Label(scrollable_frame, text="Выберите листы:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    global sheet_vars
    sheet_vars = {}
    for i, sheet_name in enumerate(workbook.sheetnames, start=1):
        if sheet_name not in {'Main data', 'Contents'}:
            var = tk.BooleanVar()
            tk.Checkbutton(scrollable_frame, text=sheet_name, variable=var).grid(row=i, column=0, padx=5, pady=2, sticky="w")
            sheet_vars[sheet_name] = var

    canvas.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox("all"))

# Загрузка файлов
word_file = "Templates/Word_templates/Act_template.docx"

main_data = [
    ({'cell_row': '1'}, {'name': '_OBJECT-NAME_', 'cell_row': '1', 'cell_column': 'B'}),
    ({'cell_row': '2'}, {'name': '_CONTRACTOR-REPRESENTATIVE_', 'cell_row': '2', 'cell_column': 'B'}),
    ({'cell_row': '3'}, {'name': '_CONTRACTOR-REPRESENTATIVE-NAME_', 'cell_row': '3', 'cell_column': 'B'}),
    ({'cell_row': '4'}, {'name': '_TECHNICAL-SUPERVISION-REPRESENTATIVE_', 'cell_row': '4', 'cell_column': 'B'}),
    ({'cell_row': '5'}, {'name': '_TECHNICAL-SUPERVISION-REPRESENTATIVE-NAME_', 'cell_row': '5', 'cell_column': 'B'}),
    ({'cell_row': '6'}, {'name': '_DESIGN-ORGANIZATION-REPRESENTATIVE_', 'cell_row': '6', 'cell_column': 'B'}),
    ({'cell_row': '7'}, {'name': '_DESIGN-ORGANIZATION-REPRESENTATIVE-NAME_', 'cell_row': '7', 'cell_column': 'B'}),
    ({'cell_row': '8'}, {'name': '_ADDITIONAL-REPRESENTATIVES_', 'cell_row': '8', 'cell_column': 'B'}),
    ({'cell_row': '9'}, {'name': '_ADDITIONAL-REPRESENTATIVES-NAME_', 'cell_row': '9', 'cell_column': 'B'}),
    ({'cell_row': '10'}, {'name': '_GENERAL-CONTRACTOR_', 'cell_row': '10', 'cell_column': 'B'})
]

subobject_data = [
    {'name': '_SUBOBJECT-NAME_', 'cell_column': 'B', 'cell_row': '1'},
    {'name': '_ACT-NUMBER_', 'cell_column': 'A'},
    {'name': '_EXECUTION-DATE-MONTH_', 'cell_column': 'B'},
    {'name': '_EXECUTION-DATE_', 'cell_column': 'B'},
    {'name': '_WORK-NAMING_', 'cell_column': 'C'},
    {'name': '_ALBUM-NAME_', 'cell_column': 'D'},
    {'name': '_PAGE_', 'cell_column': 'E'},
    {'name': '_MATERIALS_', 'cell_column': 'F'},
    {'name': '_EXECUTIVE-DIAGRAM_', 'cell_column': 'G'},
    {'name': '_LABORATORY_', 'cell_column': 'H'},
    {'name': '_END-DATE_', 'cell_column': 'I'},
    {'name': '_NEXT-WORKS_', 'cell_column': 'J'}
]

wb = None
excel_file_name = None
excel_file_path = None  # Переменная для хранения полного пути к файлу

def create_word_doc(template_path, output_path, replacements):
    doc = Document(template_path)
    for para in doc.paragraphs:
        for run in para.runs:
            for key, value in replacements.items():
                if key in run.text:
                    run.text = run.text.replace(key, value)
    doc.save(output_path)

def generate_document():
    if wb is None:
        messagebox.showwarning("Предупреждение", "Сначала загрузите файл Excel!")
        return

    selected_sheets = [sheet_name for sheet_name, var in sheet_vars.items() if var.get()]
    if not selected_sheets:
        messagebox.showwarning("Предупреждение", "Выберите хотя бы один лист!")
        return

    # Сохраняем данные перед генерацией документов
    save_to_excel()

    created_files = []
    errors = []

    base_dir = os.path.join(os.getcwd(), "Documents")
    os.makedirs(base_dir, exist_ok=True)

    excel_dir = os.path.join(base_dir, excel_file_name)
    os.makedirs(excel_dir, exist_ok=True)

    for sheet_name in selected_sheets:
        ws = wb[sheet_name]
        last_row = ws.max_row
        while last_row > 0 and all(cell.value is None for cell in ws[last_row]):
            last_row -= 1

        sheet_dir = os.path.join(excel_dir, sheet_name)
        os.makedirs(sheet_dir, exist_ok=True)

        for row_number in range(3, last_row + 1):
            replacements = {}
            for _, value_data in main_data:
                replacements[value_data['name']] = entries[value_data['name']].get()

            for obj in subobject_data:
                if obj['name'] == '_EXECUTION-DATE-MONTH_':
                    date_value = get_cell_value(ws, row_number, obj['cell_column'])
                    new_value = format_date(date_value) if date_value else ""
                    replacements[obj['name']] = new_value
                elif obj['name'] == '_SUBOBJECT-NAME_':
                    replacements[obj['name']] = get_cell_value(ws, obj['cell_row'], obj['cell_column'])
                else:
                    replacements[obj['name']] = get_cell_value(ws, row_number, obj['cell_column'])

                if obj['name'] == '_ACT-NUMBER_':
                    act_number = get_cell_value(ws, row_number, obj['cell_column']).replace('/', '_')
                    output_path = os.path.join(sheet_dir, f"{act_number}.docx")

            try:
                create_word_doc(word_file, output_path, replacements)
                created_files.append(output_path)
            except Exception as e:
                errors.append(f"Ошибка при создании {output_path}: {str(e)}")

    if created_files:
        success_message = "Успешно созданы документы:\n" + "\n".join(created_files)
        messagebox.showinfo("Успех", success_message)
    if errors:
        error_message = "Ошибки при создании документов:\n" + "\n".join(errors)
        messagebox.showerror("Ошибка", error_message)
    if not created_files and not errors:
        messagebox.showinfo("Информация", "Документы не были созданы.")

def toggle_all_sheets():
    select_all = all_var.get()
    for var in sheet_vars.values():
        var.set(select_all)

# UI Setup
root = tk.Tk()
root.title("Создание документов")

load_button = tk.Button(root, text="Загрузить Excel", command=select_excel_file)
load_button.grid(row=0, column=0, padx=5, pady=5, sticky="nw")

create_project_button = tk.Button(root, text="Создать проект", command=create_project)
create_project_button.grid(row=0, column=1, padx=5, pady=5, sticky="nw")

# Новая кнопка "Сохранить"
save_button = tk.Button(root, text="Сохранить", command=save_to_excel)
save_button.grid(row=0, column=2, padx=5, pady=5, sticky="nw")

frame = tk.Frame(root)
frame.grid(row=11, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

canvas = tk.Canvas(frame)
scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
scrollable_frame = tk.Frame(canvas)

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

all_var = tk.BooleanVar()
tk.Checkbutton(frame, text="Выбрать все", variable=all_var, command=toggle_all_sheets).grid(row=0, column=0, padx=5, pady=5, sticky="w")

canvas.grid(row=1, column=0, sticky="nsew")
scrollbar.grid(row=1, column=1, sticky="ns")

frame.grid_rowconfigure(1, weight=1)
frame.grid_columnconfigure(0, weight=1)

generate_button = tk.Button(root, text="Создать документ", command=generate_document)
generate_button.grid(row=12, column=0, columnspan=2, pady=10)

root.geometry("800x600")
root.grid_rowconfigure(11, weight=1)
root.grid_columnconfigure(1, weight=1)

sheet_vars = {}
entries = {}
excel_file_name = None
excel_file_path = None

root.mainloop()