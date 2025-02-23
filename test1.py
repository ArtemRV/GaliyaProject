import openpyxl
from docx import Document
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
import os

def format_date(date_str):
    months = {
        "01": "января", "02": "февраля", "03": "марта", "04": "апреля",
        "05": "мая", "06": "июня", "07": "июля", "08": "августа",
        "09": "сентября", "10": "октября", "11": "ноября", "12": "декабря"
    }
    
    day, month, year = date_str.split(".")
    return f"«{day}» {months[month]} {year}"

def get_cell_value(ws, cell_row, cell_column):
    cell = f"{cell_column}{cell_row}"
    return str(ws[cell].value if ws[cell].value is not None else "")

# Функция выбора файла Excel
def select_excel_file():
    global wb
    file_path = filedialog.askopenfilename(
        title="Выберите файл Excel",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        initialdir=os.path.join(os.getcwd(), "Excel_files")  # Папка Excel_files в проекте
    )
    if file_path:
        try:
            wb = openpyxl.load_workbook(file_path)
            load_ui_data(wb)
            messagebox.showinfo("Успех", f"Загружен файл: {file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить файл: {str(e)}")
    else:
        messagebox.showwarning("Предупреждение", "Файл не выбран!")

# Функция загрузки данных в интерфейс после выбора файла
def load_ui_data(workbook):
    # Обновляем поля ввода
    row0_entry.delete(0, tk.END)
    row0_entry.insert(0, get_cell_value(workbook.active, main_data[0]['cell_row'], main_data[0]['cell_column']))
    row1_entry.delete(0, tk.END)
    row1_entry.insert(0, get_cell_value(workbook.active, main_data[1]['cell_row'], main_data[1]['cell_column']))

    # Очищаем старые чекбоксы
    for widget in scrollable_frame.winfo_children():
        widget.destroy()

    # Создаем новые чекбоксы для листов
    tk.Label(scrollable_frame, text="Выберите листы:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    global sheet_vars
    sheet_vars = {}
    for i, sheet_name in enumerate(workbook.sheetnames, start=1):
        var = tk.BooleanVar()
        tk.Checkbutton(scrollable_frame, text=sheet_name, variable=var).grid(row=i, column=0, padx=5, pady=2, sticky="w")
        sheet_vars[sheet_name] = var

    # Обновляем область прокрутки
    canvas.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox("all"))

# Загрузка файлов
word_file = "template.docx"

main_data = [
    {'name': '_OBJECTNAME_', 'cell_column': 'B', 'cell_row': '1'},
    {'name': '_SUBOBJECT-NAME_', 'cell_column': 'B', 'cell_row': '2'}
]

subobject_data = [
    {'name': 'ACT_NUMBER', 'cell_column': 'A'},
    {'name': 'EXECUTION_DATE_MONTH', 'cell_column': 'B'},
    {'name': 'EXECUTION_DATE', 'cell_column': 'B'},
    {'name': 'WORK_NAMING', 'cell_column': 'C'},
    {'name': 'ALBUM_NAME', 'cell_column': 'D'},
    {'name': 'PAGE', 'cell_column': 'E'},
    {'name': 'MATERIALS', 'cell_column': 'F'},
    {'name': 'EXECUTIVE_DIAGRAM', 'cell_column': 'G'},
    {'name': 'LABORATORY', 'cell_column': 'H'},
    {'name': 'END_DATE', 'cell_column': 'I'},
    {'name': 'NEXT_WORKS', 'cell_column': 'J'}
]

# Переменная для хранения workbook
wb = None

def create_word_doc(template_path, output_path, replacements):
    doc = Document(template_path)
    for para in doc.paragraphs:
        for run in para.runs:
            for key, value in replacements.items():
                if key in run.text:
                    run.text = run.text.replace(key, value)
    doc.save(output_path)

def generate_document():
    if wb is None:
        messagebox.showwarning("Предупреждение", "Сначала загрузите файл Excel!")
        return

    selected_sheets = [sheet_name for sheet_name, var in sheet_vars.items() if var.get()]
    if not selected_sheets:
        messagebox.showwarning("Предупреждение", "Выберите хотя бы один лист!")
        return

    for sheet_name in selected_sheets:
        ws = wb[sheet_name]
        last_row = ws.max_row
        while last_row > 0 and all(cell.value is None for cell in ws[last_row]):
            last_row -= 1

        for row_number in range(4, last_row + 1):
            replacements = {}
            replacements[main_data[0]['name']] = row0_entry.get()
            replacements[main_data[1]['name']] = row1_entry.get()

            for obj in subobject_data:
                if obj['name'] == 'EXECUTION_DATE_MONTH':
                    date_value = get_cell_value(ws, row_number, obj['cell_column'])
                    new_value = format_date(date_value) if date_value else ""
                    replacements[obj['name']] = new_value
                else:
                    replacements[obj['name']] = get_cell_value(ws, row_number, obj['cell_column'])

                if obj['name'] == 'ACT_NUMBER':
                    act_number = get_cell_value(ws, row_number, obj['cell_column']).replace('/', '_')
                    output_path = f"{act_number}_{sheet_name}.docx"

            try:
                create_word_doc(word_file, output_path, replacements)
                messagebox.showinfo("Успех", f"Документ сохранен как {output_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))

def toggle_all_sheets():
    select_all = all_var.get()
    for var in sheet_vars.values():
        var.set(select_all)

# UI Setup
root = tk.Tk()
root.title("Создание документов")

# Кнопка загрузки файла
load_button = tk.Button(root, text="Загрузить Excel", command=select_excel_file)
load_button.grid(row=0, column=0, padx=5, pady=5, sticky="nw")

# Поля ввода
tk.Label(root, text="Объект:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
row0_entry = tk.Entry(root, width=100)
row0_entry.grid(row=1, column=1, padx=5, pady=5)

tk.Label(root, text="Субобъект:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
row1_entry = tk.Entry(root, width=100)
row1_entry.grid(row=2, column=1, padx=5, pady=5)

# Фрейм для списка листов
frame = tk.Frame(root)
frame.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

canvas = tk.Canvas(frame)
scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
scrollable_frame = tk.Frame(canvas)

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

all_var = tk.BooleanVar()
tk.Checkbutton(frame, text="Выбрать все", variable=all_var, command=toggle_all_sheets).grid(row=0, column=0, padx=5, pady=5, sticky="w")

canvas.grid(row=1, column=0, sticky="nsew")
scrollbar.grid(row=1, column=1, sticky="ns")

frame.grid_rowconfigure(1, weight=1)
frame.grid_columnconfigure(0, weight=1)

# Кнопка генерации
generate_button = tk.Button(root, text="Создать документ", command=generate_document)
generate_button.grid(row=4, column=0, columnspan=2, pady=10)

# Настройка окна
root.geometry("800x400")
root.grid_rowconfigure(3, weight=1)
root.grid_columnconfigure(1, weight=1)

# Инициализация пустого списка листов
sheet_vars = {}

root.mainloop()